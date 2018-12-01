# -*- coding: utf-8 -*-
"""
Created on Tue Nov 27 07:05:57 2018

@author: matte
"""

import os
import pyautogui
import subprocess

import requests
import openpyxl
from bs4 import BeautifulSoup

wb = openpyxl.load_workbook('words.xlsx')
sheet = wb.get_sheet_by_name('vocabulary')
sheet.cell(row=1, column=1).value

#put in cell range here
for i in range(1,10):
    searchCont = sheet.cell(row=i, column=1).value

    response = requests.get('https://en.oxforddictionaries.com/definition/' + searchCont)
    soup = BeautifulSoup(response.text, 'html.parser')
    engDefs = soup.find(class_='ind').get_text()  
    
    response2 = requests.get('https://www.thesaurus.com/browse/' + searchCont)
    soup = BeautifulSoup(response2.text, 'html.parser')   
    engThes = soup.find(class_='css-161w7kd eqpevqj3').get_text()
    
    os.startfile(r'C:\Program Files (x86)\Anki\anki.exe') 
    pyautogui.PAUSE = 1
    pyautogui.press('a')
    pyautogui.press('a')
    pyautogui.press('tab')
    pyautogui.typewrite(searchCont)    
    pyautogui.press('tab')
    pyautogui.typewrite(engThes)   
    pyautogui.press('tab')
    pyautogui.typewrite(engDefs)
    pyautogui.press('tab')
    pyautogui.typewrite('tag')
    pyautogui.press('tab')
    pyautogui.press('enter')
    
    
    #print (searchCont)
    #print (engThes)
    #print (engDefs)
    #print (' ')a
    
    
    
